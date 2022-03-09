import time

import openpyxl
from selenium import webdriver

excel_path="D:\check\Excel\kk.xlsx"

open_workbook=openpyxl.load_workbook(excel_path)
open_worksheet=open_workbook["Test Case"]
max_row=open_worksheet.max_row
max_column=open_worksheet.max_column
print(max_row,max_column)
print(open_worksheet.cell(2,2).value)
for r in range(2,max_row+1):
    print(r)
    print(open_worksheet.cell(r,6).value)

open_worksheet.cell(16,3).value="test"
open_workbook.save("D:\check\Excel\kk1.xlsx")
open_workbook.close()
driver=webdriver.Chrome(executable_path="C:\\Users\Karun\\PycharmProjects\\Sel;enium with Python\\Drivers\\chromedriver.exe")
driver.maximize_window()
driver.implicitly_wait(10)
driver.get("http://192.168.0.199:9091/QuaLISWeb/#/login")
driver.find_element_by_xpath("//*[@name='sloginid']").send_keys("cdolman")
time.sleep(3)
driver.find_element_by_xpath("//*[@name='spassword']").send_keys("12345")
time.sleep(3)
driver.find_element_by_xpath("//*[text()='Login']").click()
time.sleep(2)
toast_message=driver.find_element_by_xpath("//*[@class='Toastify']").text
print(toast_message)

