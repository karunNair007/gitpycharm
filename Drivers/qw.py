

import openpyxl


excel_path="D:\check\Excel\kk1.xlsx"

open_workbook=openpyxl.load_workbook(excel_path)
open_worksheet=open_workbook["Test Case"]
max_row=open_worksheet.max_row
max_column=open_worksheet.max_column
print(max_row,max_column)
#print(open_worksheet.cell(2,2).value)
kkk=[]
for i in range(2,max_row+1):
    kk = []
    #print(i)
    for j in range(1,max_column+1):
        #print(open_worksheet.cell(i, j).value)


        hh=open_worksheet.cell(i, j).value
        kk.insert(j,hh)
    print(kk)
    kkk.insert(i,kk)
    print(kkk)
