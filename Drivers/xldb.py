import openpyxl
import pypyodbc

databasecredentials="Driver={SQL Server};Server=192.168.0.199\SQLEXPRESS2019;Database=NIBSCNEW-01-02-2022;UID=sa;PWD=Krishna@007"
conn = pypyodbc.connect(databasecredentials)
print(conn)
cursor1 = conn.cursor()
updatescript="update manufacturercontactinfo set semail='xxxx1@agaramtech.com'"
script=cursor1.execute(updatescript)
selectscript="select * from manufacturercontactinfo "
script2=cursor1.execute(selectscript)
excel_path="D:\check\Excel\kk.xlsx"

open_workbook=openpyxl.load_workbook(excel_path)
open_worksheet=open_workbook["Test Case"]
max_row=open_worksheet.max_row
#print(open_worksheet.cell(2,2).value)
ROW=COLUMN=1

kk=[ i for i in script2]
print(kk)
print(len(kk))
for j  in range(len(kk)):
    print(kk[j])
    s=str(kk[j])
    s1=s.split(",")
    print(s1)
    for k in range(len(s1)):
        print(s1[k].strip())
        s2=str(s1[k])
        open_worksheet.cell(ROW,COLUMN).value=s2
        COLUMN=COLUMN+1
        v=')' in s2
        s3=s2.strip()

        v1=s3.find(')')
        print("find value",v1)
        if v1==1 or v1==2:
            print(") is present")
            ROW+=1
            COLUMN=1
        open_workbook.save(excel_path)
        open_workbook.close()

