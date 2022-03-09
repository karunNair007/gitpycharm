import openpyxl


def jj():
    excel_path = "D:\check\Excel\kk.xlsx"

    open_workbook = openpyxl.load_workbook(excel_path)
    open_worksheet = open_workbook["Test Case"]
    max_row = open_worksheet.max_row
    max_column = open_worksheet.max_column
    kkk = []

    print(max_row, max_column)
    for i in range(2, max_row + 1):
        kk = []
        for j in range(1, max_column + 1):
            data = open_worksheet.cell(i, j).value
            kk.insert(j, data)

        # print(kk)
        kkk.insert(i, kk)
    return kkk



print(jj())

