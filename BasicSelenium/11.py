import openpyxl
import pypyodbc

databasecredentials="Driver={SQL Server};Server=192.168.0.199\SQLEXPRESS2019;Database=NIBSCNEW-01-02-2022;UID=sa;PWD=Krishna@007"
conn = pypyodbc.connect(databasecredentials)
print(conn)
cursor1 = conn.cursor()
updatescript="update manufacturercontactinfo set semail='xxxx1@agaramtech.com'"
script=cursor1.execute(updatescript)
selectscript="select * from manufacturercontactinfo "
script2=cursor1.execute(selectscript)
excel_path="D:\check\Excel\kk.xlsx"

open_workbook=openpyxl.load_workbook(excel_path)
open_worksheet=open_workbook["Test Case"]
max_row=open_worksheet.max_row
#print(open_worksheet.cell(2,2).value)
row=1

for i in script2:
    column=1
    print(i)
    for j in i:
        print(j)
        open_worksheet.cell(row, column).value = j
        column += 1
    row+=1
    open_workbook.save(excel_path)
    open_workbook.close()










